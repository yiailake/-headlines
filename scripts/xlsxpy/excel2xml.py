from openpyxl import load_workbook

wb = load_workbook('C:/Users/Administrator/Desktop/xlsxpy/strings.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']

##i = 0
##for column in sheet.columns:
##    for cell in column:
##        if cell.value == None:
##            continue
##        else:
##            #print(cell.value)
##            clist = []
##            clist.append(str(cell.value) + '\n')
##            i = i + 1
##            path = "C:/Users/Administrator/Desktop/test/" + str(i) + ".txt"
##            with open(path, 'w', encoding='UTF-8') as f:
##                 f.writelines(clist)
##                 f.close()

print(len(list(sheet.columns)))

for i in range(0, 12):
    clist = []
    for cell in list(sheet.columns)[i]:
        if cell.value == None:
            continue
        else:
            clist.append(str(cell.value) + '\n')
    path = "C:/Users/Administrator/Desktop/xlsxpy/xml/" + str(i) + ".xml"
    f = open(path, "w", encoding='UTF-8')
    f.writelines(clist)
    f.close()


    




